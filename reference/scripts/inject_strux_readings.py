#!/usr/bin/env python3
"""Append synthetic readings for STRUX-only meters to a site dataset.

Some Excel-referenced meters (e.g. B611.T4-A3, T4-C1, T4-C4) have no Snowflake
timeseries — their monthly values live in the STRUX tab of the source xlsx.
Without these readings the ontology can't compute per-building consumption for
the buildings those meters are accounted against, producing large topology-vs-
Excel diffs that are actually just missing data.

This helper reads STRUX monthly values and emits synthetic counter readings
(one per month-end, cumulative) into the site's readings.csv for any
timeseries_ref whose external_id appears in STRUX but has no readings yet.

Usage:
    python inject_strux_readings.py data/sites/gartuna reference/monthly_reporting_documents/inputs/gtn.xlsx
"""

from __future__ import annotations

import csv
import sys
from calendar import monthrange
from pathlib import Path

from openpyxl import load_workbook


def load_strux_monthly(xlsx_path: Path) -> dict[str, dict[str, float]]:
    """Return {meter_id: {'YYYY-MM': monthly_value, ...}} from the STRUX tab."""
    wb = load_workbook(xlsx_path, data_only=True)
    if "STRUX" not in wb.sheetnames:
        return {}
    ws = wb["STRUX"]
    result: dict[str, dict[str, float]] = {}
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), start=1):
        if row and row[3] and str(row[3]).lower().startswith("mätarbeteckning"):
            header_row = i
            break
    if header_row is None:
        return {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        mid = row[3]
        if mid is None:
            continue
        mid_str = str(mid)
        monthly: dict[str, float] = {}
        for col_idx in range(12):
            val = row[7 + col_idx]
            if isinstance(val, (int, float)):
                monthly[f"2026-{col_idx + 1:02d}"] = float(val)
        if monthly:
            result[mid_str] = monthly
    return result


def main() -> int:
    if len(sys.argv) != 3:
        print(__doc__, file=sys.stderr)
        return 2
    site_dir = Path(sys.argv[1])
    xlsx_path = Path(sys.argv[2])

    ts_refs_path = site_dir / "timeseries_refs.csv"
    readings_path = site_dir / "readings.csv"
    refs = list(csv.DictReader(open(ts_refs_path)))

    existing_ts = set()
    with open(readings_path) as f:
        for r in csv.DictReader(f):
            existing_ts.add(r["timeseries_id"])

    strux = load_strux_monthly(xlsx_path)

    new_rows: list[dict[str, str]] = []
    for tr in refs:
        if tr.get("kind") != "raw":
            continue
        ts_id = tr["timeseries_id"]
        if ts_id in existing_ts:
            continue
        ext_id = tr.get("external_id", "")
        monthly = strux.get(ext_id)
        if not monthly:
            continue
        # Emit cumulative counter readings: anchor 2025-12-31=0, then end-of-month
        cumulative = 0.0
        new_rows.append({
            "timeseries_id": ts_id,
            "timestamp": "2025-12-31",
            "value": "0.0",
            "recorded_at": "",
        })
        for month in sorted(monthly):
            year, m = map(int, month.split("-"))
            last_day = monthrange(year, m)[1]
            cumulative += monthly[month]
            new_rows.append({
                "timeseries_id": ts_id,
                "timestamp": f"{year}-{m:02d}-{last_day:02d}",
                "value": str(cumulative),
                "recorded_at": "",
            })
        print(f"injecting {len(monthly)} monthly readings for {ts_id} (external_id={ext_id})")

    if not new_rows:
        print("nothing to inject")
        return 0

    with open(readings_path, "a", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["timeseries_id", "timestamp", "value", "recorded_at"])
        for r in new_rows:
            w.writerow(r)
    print(f"appended {len(new_rows)} rows to {readings_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
