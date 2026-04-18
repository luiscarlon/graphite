#!/usr/bin/env python3
"""Spot-check per-building monthly totals: evaluate each Excel formula using
Snowflake-derived monthly deltas, compare to the Excel's reported building
totals. Supports recursive virtual meters (a building that appears as both
an accounting row AND as a `+` / `−` term in another building's formula).

Output:
  - ``04_validation/building_totals_spot_check.csv`` — one row per
    (building, month): excel value, topology-derived value, Δ, Δ%, any
    unresolved term names.
  - console summary of totals + largest divergences.

Usage:
    python validate_building_totals.py WORKSTREAM_DIR \\
        --excel-sheet Ånga --months 2026-01,2026-02
"""

from __future__ import annotations

import argparse
import collections
import csv
import re
import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")
from openpyxl import load_workbook  # noqa: E402


def canon(m: str) -> str:
    """Canonicalise across naming conventions:
    - dash-separated (kyla `B600-KB2`) → dot+underscore (`B600.KB2`)
    - strip trailing `_E` energy-variant suffix
    - unify `VM##` → `VMM##`
    """
    x = m
    if "-" in x and "." not in x:
        parts = x.split("-")
        x = parts[0] + "." + "_".join(parts[1:])
    x = re.sub(r"_E$", "", x)
    x = re.sub(r"\.(\w+?)_V(M)(\d+)$", r".\1_VMM\3", x)
    return x


def load_crosswalk(path: Path) -> dict[str, str]:
    """Return canonical_id → snowflake_id lookup."""
    out: dict[str, str] = {}
    if not path.exists():
        return out
    with path.open() as f:
        for r in csv.DictReader(f):
            snow = r.get("snowflake_id") or ""
            if snow:
                out[r["facit_id"]] = snow
                # also map the excel_label to its snowflake_id
                ex = r.get("excel_label") or ""
                if ex:
                    out[ex] = snow
                    out[canon(ex)] = snow
    return out


def load_monthly_ts(path: Path) -> dict[tuple[str, str], float]:
    out: dict[tuple[str, str], float] = {}
    if not path.exists():
        return out
    with path.open() as f:
        for r in csv.DictReader(f):
            out[(r["meter_id"], r["month"])] = float(r["delta"])
    return out


def load_formulas(path: Path) -> dict[str, dict]:
    """Return {building: {'add': [(meter_raw, faktor)], 'sub': [(meter_raw, faktor)]}}."""
    out: dict[str, dict] = collections.defaultdict(lambda: {"add": [], "sub": []})
    with path.open() as f:
        for r in csv.DictReader(f):
            fs = r.get("faktor") or ""
            try:
                fak = float(fs) if fs else 1.0
            except ValueError:
                fak = 1.0
            out[r["building"]][r["role"]].append((r["meter_id"], fak))
    return out


def load_excel_totals(xlsx: Path, sheet: str, months: list[str]) -> dict[str, dict[str, float]]:
    """Return {building: {month: value}} from the Excel sheet.

    Sheet convention (värme/ånga/kyla all share it):
    - row 7: header — col C = Jan, D = Feb, etc.
    - rows 8+: per-building. col B = building number.
    """
    wb = load_workbook(xlsx, data_only=True)
    ws = wb[sheet]
    # Build column-letter → month map from row 7 headers
    month_col: dict[str, int] = {}
    for c in range(3, ws.max_column + 1):
        v = ws.cell(7, c).value
        if hasattr(v, "year"):
            key = f"{v.year:04d}-{v.month:02d}"
            if key in months:
                month_col[key] = c

    out: dict[str, dict[str, float]] = {}
    for r in range(8, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if b is None:
            continue
        b_key = str(b).strip()
        out[b_key] = {}
        for month, col in month_col.items():
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                out[b_key][month] = float(v)
    return out


class Evaluator:
    """Evaluates Excel formula rows against monthly timeseries, supporting
    recursive virtual meters (e.g. kyla's B600-KB2 and Prod-600)."""

    def __init__(
        self,
        formulas: dict[str, dict],
        crosswalk: dict[str, str],
        ts: dict[tuple[str, str], float],
    ):
        self.formulas = formulas
        self.crosswalk = crosswalk
        self.ts = ts
        # Index formulas by canonical building name for recursive lookup
        self.by_canon = {canon(b): b for b in formulas}

    def _raw_delta(self, snowflake_id: str, month: str) -> float | None:
        return self.ts.get((snowflake_id, month))

    def _resolve_meter(self, raw_name: str, month: str, depth: int = 0) -> tuple[float | None, str | None]:
        """Return (value, error_note). Tries:
        1. direct crosswalk → timeseries
        2. canonical form → timeseries
        3. Excel formula row with same name → recurse
        """
        if depth > 8:
            return None, f"recursion too deep at {raw_name}"
        # Direct lookup
        snow = self.crosswalk.get(raw_name) or self.crosswalk.get(canon(raw_name))
        if snow:
            v = self._raw_delta(snow, month)
            if v is not None:
                return v, None
        # Try as a formula-building
        canonical = canon(raw_name)
        building_name = self.by_canon.get(canonical) or (raw_name if raw_name in self.formulas else None)
        if building_name:
            return self.evaluate(building_name, month, depth + 1)
        return None, raw_name

    def evaluate(self, building: str, month: str, depth: int = 0) -> tuple[float | None, str | None]:
        """Return (value, error) where error names the first unresolved term."""
        if building not in self.formulas:
            return None, f"no formula for {building}"
        f = self.formulas[building]
        total = 0.0
        for sign, terms in (("+", f["add"]), ("-", f["sub"])):
            for meter, fak in terms:
                v, err = self._resolve_meter(meter, month, depth + 1)
                if v is None:
                    return None, err or meter
                total += (v * fak) if sign == "+" else -(v * fak)
        return total, None


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workstream_dir", type=Path)
    ap.add_argument("--excel-sheet", required=True, help="Ånga | Värme | Kyla | ...")
    ap.add_argument("--months", default="2026-01,2026-02",
                    help="Comma-separated YYYY-MM list to check")
    ap.add_argument("--threshold-abs", type=float, default=5.0,
                    help="Flag rows with |Δ| > this (MWh)")
    ap.add_argument("--threshold-pct", type=float, default=10.0,
                    help="Flag rows with |Δ%%| > this")
    args = ap.parse_args()

    ws: Path = args.workstream_dir
    months = [m.strip() for m in args.months.split(",") if m.strip()]

    xlsx = ws / "00_inputs" / "excel_source.xlsx"
    formulas = load_formulas(ws / "01_extracted" / "excel_formulas.csv")
    crosswalk = load_crosswalk(ws / "02_crosswalk" / "meter_id_map.csv")
    ts = load_monthly_ts(ws / "01_extracted" / "timeseries_monthly.csv")
    excel = load_excel_totals(xlsx, args.excel_sheet, months)

    ev = Evaluator(formulas, crosswalk, ts)

    out_rows: list[dict] = []
    for b in sorted(formulas.keys()):
        for month in months:
            e_val = excel.get(b, {}).get(month)
            topo, err = ev.evaluate(b, month)
            row = {
                "building": b,
                "month": month,
                "excel": e_val if e_val is not None else "",
                "topology": f"{topo:.4f}" if topo is not None else "",
                "delta": f"{topo - e_val:.4f}" if (e_val is not None and topo is not None) else "",
                "delta_pct": (f"{(topo - e_val)/e_val*100:.2f}" if (e_val is not None and topo is not None and abs(e_val) > 0.1) else ""),
                "unresolved": err or "",
            }
            out_rows.append(row)

    out_path = ws / "04_validation" / "building_totals_spot_check.csv"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    cols = ["building", "month", "excel", "topology", "delta", "delta_pct", "unresolved"]
    with out_path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in out_rows:
            w.writerow(r)
    print(f"wrote {out_path} ({len(out_rows)} rows)")

    # Console summary
    print(f"\n{'building':16} {'month':8} {'excel':>10} {'topology':>10} {'Δ':>10} {'Δ%':>8}  note")
    print("-" * 90)
    tot = collections.defaultdict(lambda: [0.0, 0.0])
    flagged = 0
    unresolved = 0
    for r in out_rows:
        if r["unresolved"]:
            unresolved += 1
            print(f"  {r['building']:14} {r['month']:8} {'?':>10} {'?':>10} {'?':>10} {'?':>8}  unresolved: {r['unresolved']}")
            continue
        e = float(r["excel"]) if r["excel"] != "" else None
        t = float(r["topology"]) if r["topology"] != "" else None
        d = t - e if (e is not None and t is not None) else None
        p = (d / e * 100) if (e is not None and d is not None and abs(e) > 0.1) else None
        if e is not None and t is not None:
            tot[r["month"]][0] += e; tot[r["month"]][1] += t
        is_flag = d is not None and abs(d) > args.threshold_abs and (p is None or abs(p) > args.threshold_pct)
        if is_flag: flagged += 1
        mk = "⚠" if is_flag else " "
        e_s = f"{e:>10.2f}" if e is not None else f"{'n/a':>10}"
        t_s = f"{t:>10.2f}" if t is not None else f"{'n/a':>10}"
        d_s = f"{d:>+10.2f}" if d is not None else f"{'':>10}"
        p_s = f"{p:+.1f}%" if p is not None else "  —  "
        print(f"{mk} {r['building']:14} {r['month']:8} {e_s} {t_s} {d_s} {p_s:>8}")

    print(f"\ntotals (resolvable rows only):")
    for m in sorted(tot):
        e, t = tot[m]
        d = t - e
        p = (d / e * 100) if abs(e) > 0.1 else 0
        print(f"  {m}: excel={e:>10.2f}  topo={t:>10.2f}  Δ={d:+.2f} ({p:+.2f}%)")
    print(f"\n{unresolved} unresolved rows, {flagged} flagged "
          f"(|Δ|>{args.threshold_abs} and |Δ%|>{args.threshold_pct}%)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
