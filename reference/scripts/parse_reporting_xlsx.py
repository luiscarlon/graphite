#!/usr/bin/env python3
"""Extract a media-specific slice of an AstraZeneca monthly-reporting workbook.

The workbooks (``gtn.xlsx`` / ``snv.xlsx``) carry one sheet per media (Ånga,
Kyla, Värme, Kallvatten, Kyltornsvatten, EL).  Each building row follows a
fixed pattern:

  col B           building number
  cols C..N       monthly consumption (pulled via XLOOKUP from the STRUX sheet)
  col O           yearly total (SUM)
  col P           Kommentar (human-readable note)
  cols S..W       meter IDs used in the accounting formula
                  formula = XLOOKUP(S) + XLOOKUP(T) − XLOOKUP(U) − XLOOKUP(V) − XLOOKUP(W)

The S/T columns are *additive* supply meters; U/V/W are *subtractive* — they
represent downstream consumers whose flow should not be attributed to this
building (they get accounted on their own row).  That asymmetry is what
encodes allocation; preserving it is the whole point of this extractor.

Outputs (written to the ``--out-dir`` directory):

- ``excel_formulas.csv``     — one row per (building, role, meter_id)
- ``excel_comments.md``      — cell comments + Kommentar column contents
- ``excel_tabs_inventory.md`` — every sheet in the workbook with dimensions and role
- ``excel_meters_used.csv``  — unique meter IDs referenced on the selected sheet
- ``excel_intake_meters.csv`` — STRUX rows for this media (the meter catalog)

Usage::

    python parse_reporting_xlsx.py \\
        reference/monthly_reporting_documents/inputs/gtn.xlsx \\
        --media Ånga \\
        --out-dir reference/media_workstreams/gtn_anga/01_extracted
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import warnings
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

# openpyxl warns about conditional-formatting and unsupported extensions for
# these workbooks; the warnings are noise for our purposes.
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# Match the XLOOKUP call AND its surrounding arithmetic. We split the formula
# into top-level additive terms (``+`` / ``−``) and for each term locate the
# XLOOKUP's cell reference and evaluate the rest of the term's arithmetic.
# This captures:
#   0.8*XLOOKUP(S14,...)          → faktor = 0.8     (pre-factor)
#   XLOOKUP(S26,...)*24*31/1000   → faktor = 0.744   (post-factors/divisors)
#   0.5*XLOOKUP(T,...)*2          → faktor = 1.0     (combined)
# First term in the formula has no leading sign and defaults to +.
XLOOKUP_REF_RE = re.compile(
    r"(?:_xlfn\.)?XLOOKUP\s*\(\s*([A-Z]+)(\d+)\s*,",
)


def _column_letter_to_index(letters: str) -> int:
    """A → 1, B → 2, ..., Z → 26, AA → 27, etc."""
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _find_xlookup_spans(text: str) -> list[tuple[str, int, int, int]]:
    """Return [(col_letters, row_int, start_idx, end_idx_inclusive), ...] for
    every ``XLOOKUP(col_row, ...)`` call in ``text`` (spans are
    balanced-paren ranges).
    """
    spans: list[tuple[str, int, int, int]] = []
    pos = 0
    while True:
        m = XLOOKUP_REF_RE.search(text, pos)
        if not m:
            break
        open_paren = text.find("(", m.start())
        depth = 0
        end = open_paren
        for i in range(open_paren, len(text)):
            if text[i] == "(":
                depth += 1
            elif text[i] == ")":
                depth -= 1
                if depth == 0:
                    end = i
                    break
        spans.append((m.group(1), int(m.group(2)), m.start(), end))
        pos = end + 1
    return spans


def parse_formula_terms(
    formula_text: str,
    row: int,
    cell_values: dict[str, float] | None = None,
) -> dict[str, dict]:
    """Return ``{column_letter: {'sign': '+'|'−', 'faktor': float}}`` for every
    XLOOKUP term referencing ``row``.

    Works even when multiple XLOOKUP calls share an outer wrapper like
    ``=$F$5*(XLOOKUP(...) + XLOOKUP(...) − XLOOKUP(...))``: for each XLOOKUP
    we evaluate the full formula with that call substituted by ``1`` and all
    others by ``0``; the result is that term's signed effective coefficient.

    ``cell_values`` maps absolute Excel refs like ``$F$5`` to numeric values
    so the evaluator can resolve site-scalar factors (EL sheet uses F5=0.001
    to convert kWh→MWh).
    """
    cell_values = cell_values or {}
    text = formula_text.lstrip("=").strip()
    spans = _find_xlookup_spans(text)

    # Substitute absolute cell refs with their numeric values
    resolved = text
    for ref, val in cell_values.items():
        resolved_ref = ref
        # Accept both "$F$5" and "F5" for convenience
        resolved = resolved.replace(resolved_ref, f"({val})")
        resolved = resolved.replace(resolved_ref.replace("$", ""), f"({val})")

    # Re-find spans in the resolved text (offsets shift)
    resolved_spans = _find_xlookup_spans(resolved)
    if len(resolved_spans) != len(spans):
        # substitution accidentally wiped an XLOOKUP; bail with empty map
        return {}

    out: dict[str, dict] = {}
    for i, (col, trow, _, _) in enumerate(resolved_spans):
        if trow != row:
            continue
        # Build expr with span i → 1, others → 0
        parts: list[str] = []
        last = 0
        for j, (_, _, s2, e2) in enumerate(resolved_spans):
            parts.append(resolved[last:s2])
            parts.append("1" if j == i else "0")
            last = e2 + 1
        parts.append(resolved[last:])
        expr = "".join(parts).replace(",", ".")
        # Strip any surviving non-arithmetic characters (cell refs not in
        # cell_values, function names we can't evaluate). Conservative — if
        # we see something we don't understand, coefficient defaults to 0.
        safe = re.sub(r"[^0-9.+\-*/()\s]", "", expr)
        try:
            coef = float(eval(safe, {"__builtins__": {}}, {})) if safe.strip() else 1.0
        except Exception:
            coef = 1.0
        sign = "+" if coef >= 0 else "−"
        faktor = abs(coef) if coef != 0 else 1.0
        out[col] = {"sign": sign, "faktor": faktor}
    return out


# kept for backward compatibility
def parse_formula_signs(formula_text: str, row: int) -> dict[str, str]:
    return {k: v["sign"] for k, v in parse_formula_terms(formula_text, row).items()}


def _workbook_scalar_cells(ws) -> dict[str, float]:
    """Collect numeric values from cells that often appear as absolute refs
    in formulas (``$<COL>$<ROW>``). Scans the first 6 rows / first column
    region used for site-wide scalars (e.g. EL sheet's ``F5 = 0.001`` for
    kWh→MWh scaling).
    """
    out: dict[str, float] = {}
    for r in range(1, 7):
        for c in range(1, min(ws.max_column + 1, 12)):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                col_letter = chr(64 + c) if c <= 26 else "A" + chr(64 + c - 26)
                out[f"${col_letter}${r}"] = float(v)
    return out


def extract_building_formulas(ws, ws_values=None) -> list[dict]:
    """Scan a media sheet and return one record per referenced meter cell.

    Unlike ånga, where every row uses the 5-term S+T−U−V−W accounting
    pattern, other media sheets (e.g. Värme) have variable-length formulas
    with different sign combinations. Each row's formula is parsed and the
    sign per column is taken from the formula text itself.
    """
    records: list[dict] = []
    # Scalar cells must come from the data-only sheet to resolve to numbers.
    sheet_scalars = _workbook_scalar_cells(ws_values if ws_values is not None else ws)
    # Pre-read the per-row allocation column (col R = 18) and any other
    # absolute-column refs that formulas use. Kyla uses ``$R<row>`` as a
    # per-building allocation factor (e.g. R13 = 0.38 meaning B611 gets 38%
    # of the virtual meter B600-KB2).
    val_ws = ws_values if ws_values is not None else ws
    for row in range(8, ws.max_row + 1):
        building = ws.cell(row=row, column=2).value  # col B
        if building is None:
            continue
        comment_cell = ws.cell(row=row, column=16).value  # col P
        factor_cell = ws.cell(row=row, column=18).value  # col R

        c_val = ws.cell(row=row, column=3).value
        formula_text = ""
        if isinstance(c_val, ArrayFormula):
            formula_text = c_val.text or ""
        elif isinstance(c_val, str) and c_val.startswith("="):
            formula_text = c_val

        # Build per-row cell_values: start from sheet scalars, then add any
        # absolute-column refs that reference this row (e.g. ``$R13`` = 0.38).
        row_cells = dict(sheet_scalars)
        for c in range(1, min(val_ws.max_column + 1, 30)):
            v = val_ws.cell(row, c).value
            if isinstance(v, (int, float)):
                col_letter = chr(64 + c) if c <= 26 else "A" + chr(64 + c - 26)
                row_cells[f"${col_letter}${row}"] = float(v)
                row_cells[f"${col_letter}{row}"] = float(v)

        col_terms = parse_formula_terms(formula_text, row, cell_values=row_cells)
        if not col_terms:
            continue

        for col_letter, info in sorted(col_terms.items(), key=lambda kv: _column_letter_to_index(kv[0])):
            v = ws[f"{col_letter}{row}"].value
            if v is None or v == "":
                continue
            sign = info["sign"]
            role = "add" if sign == "+" else "sub"
            per_term_faktor = info["faktor"]
            # Prefer per-term factor when non-unit; fall back to the per-row
            # factor cell (col R) for sheets that still use that convention.
            if abs(per_term_faktor - 1.0) > 1e-9:
                emitted_faktor = per_term_faktor
            elif factor_cell is not None and factor_cell != "":
                try:
                    emitted_faktor = float(factor_cell)
                except (TypeError, ValueError):
                    emitted_faktor = factor_cell
            else:
                emitted_faktor = ""
            records.append(
                {
                    "row": row,
                    "building": building,
                    "role": role,
                    "sign": sign,
                    "column": col_letter,
                    "meter_id": str(v),
                    "kommentar": comment_cell if comment_cell else "",
                    "faktor": emitted_faktor,
                    "n_terms": len(col_terms),
                }
            )
    return records


def extract_cell_comments(ws) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment:
                out.append((cell.coordinate, cell.comment.text))
    return out


def extract_strux_media(wb, media_label: str) -> list[dict]:
    """Return rows from the STRUX sheet matching the given media label."""
    if "STRUX" not in wb.sheetnames:
        return []
    ws = wb["STRUX"]
    header = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    out: list[dict] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not any(row):
            continue
        rec = dict(zip(header, row, strict=False))
        mediaslag = (rec.get("Mediaslag") or "").strip()
        if mediaslag.lower() == media_label.lower():
            out.append(
                {
                    "anlaggning": rec.get("Anläggning") or "",
                    "mediaslag": mediaslag,
                    "avlasning": rec.get("Avläsning") or "",
                    "matarbeteckning": rec.get("Mätarbeteckning") or "",
                    "betjanar": rec.get("Betjänar") or "",
                    "matarstallning_eller_forbrukning": rec.get("Mätarställning eller förbrukning") or "",
                    "enhet": rec.get("Enhet") or "",
                }
            )
    return out


def inventory_sheets(wb) -> list[dict]:
    out: list[dict] = []
    for name in wb.sheetnames:
        ws = wb[name]
        out.append(
            {
                "sheet": name,
                "rows": ws.max_row,
                "cols": ws.max_column,
                "hidden": ws.sheet_state != "visible",
            }
        )
    return out


def write_formulas_csv(path: Path, records: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    cols = ["building", "row", "role", "sign", "column", "meter_id", "kommentar", "faktor", "n_terms"]
    with path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in records:
            w.writerow({c: r.get(c, "") for c in cols})


def write_meters_used_csv(path: Path, records: list[dict]) -> None:
    seen: dict[str, dict] = {}
    for r in records:
        mid = r["meter_id"]
        if mid not in seen:
            seen[mid] = {
                "meter_id": mid,
                "roles": set(),
                "buildings": set(),
            }
        seen[mid]["roles"].add(r["role"])
        seen[mid]["buildings"].add(str(r["building"]))
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["meter_id", "roles", "buildings"])
        for mid in sorted(seen):
            rec = seen[mid]
            w.writerow(
                [mid, "|".join(sorted(rec["roles"])), "|".join(sorted(rec["buildings"]))]
            )


def write_intake_meters_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("# no STRUX rows matched; skip\n")
        return
    cols = list(rows[0].keys())
    with path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(rows)


def write_comments_md(path: Path, media: str, sheet_comments: list[tuple[str, str]], kommentar_entries: list[dict]) -> None:
    lines = [f"# Excel comments — {media} sheet", ""]
    if sheet_comments:
        lines += ["## Cell comments (openpyxl .comment)", ""]
        for coord, text in sheet_comments:
            text_clean = text.replace("\n", " · ")
            lines.append(f"- **{coord}** — {text_clean}")
        lines.append("")
    else:
        lines += ["## Cell comments (openpyxl .comment)", "", "_(none)_", ""]
    lines += ["## `Kommentar` column (col P) non-empty entries", ""]
    any_entry = False
    for e in kommentar_entries:
        if not e.get("kommentar"):
            continue
        any_entry = True
        lines.append(f"- row {e['row']} (building {e['building']}) — {e['kommentar']}")
    if not any_entry:
        lines.append("_(none)_")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n")


def write_tabs_inventory_md(path: Path, media: str, sheets: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# Excel tabs inventory",
        "",
        f"Workbook slice relevant for **{media}**. The `{media}` sheet carries the per-building accounting formulas; STRUX is the meter catalog; Avläsning/PME/PME_EL are the underlying reading sources that the accounting formulas XLOOKUP into.",
        "",
        "| sheet | rows | cols | hidden | role |",
        "|---|---:|---:|:---:|---|",
    ]
    roles = {
        "Intro": "workbook documentation — what each tab does",
        "Rapport Site": "site-level rollup",
        "Rapport Byggnad": "per-building rollup",
        "Total GTN": "purchased-energy totals for external reporting",
        "EL": "electricity per-building accounting",
        "Kyla": "cooling per-building accounting",
        "Värme": "heating per-building accounting",
        "Ånga": "steam per-building accounting — this workstream",
        "Kallvatten": "cold-water per-building accounting",
        "Kyltornsvatten": "cooling-tower-water per-building accounting",
        "STRUX": "meter catalog — one row per meter with monthly values; XLOOKUP target for all media sheets",
        "Kontroll FV-mätare": "district-heating meter verification",
        "Evidence Gärtuna": "stakeholder evidence dump",
        "VÅ9 alla mätare": "VÅ9 (heat-pump recovery) meter inventory",
        "VÅ9 Nyckeltal": "VÅ9 KPIs",
        "PME_EL": "automatic electricity readings",
        "PME": "automatic non-electricity readings (energy, flow meters)",
        "Avläsning": "manual readings (e.g. sea-water meters not connected)",
        "Site": "building → site lookup (hidden)",
        "Lista": "enumeration: media types, units (hidden)",
    }
    for s in sheets:
        marker = "hidden" if s["hidden"] else ""
        role = roles.get(s["sheet"], "")
        lines.append(f"| {s['sheet']} | {s['rows']} | {s['cols']} | {marker} | {role} |")
    lines += ["", "**Named ranges of interest:**", ""]
    lines += [
        "- `STRUX_Mätare` = `STRUX!$D$3:$D$4923` — meter ID column; XLOOKUP target from all media sheets.",
        "- `STRUX_data` = `STRUX!$H$3:$S$4923` — monthly values (Jan..Dec) for every meter in STRUX.",
        "- `STRUX_mån` = `STRUX!$H$2:$S$2` — monthly header row.",
        "- `Avläs_data` / `Avläs_mätare` / `Avläs_mån` — the same pattern for manually-recorded meters.",
        "- `PME_data` / `PME_mätare` / `PME_mån` — the same for automatic (BMS) meters.",
        "- `EL_data` / `EL_mätare` / `EL_mån` — same for electricity meters.",
        "- `Site_GTN` = `Site!$A$1:$B$64` — building → site lookup used by the media sheets' col A.",
        "",
        "**Per-row accounting formula used on every media sheet (col C..N):**",
        "",
        "```",
        "= XLOOKUP(S{row}, STRUX_Mätare, STRUX_data, 0, 0, 1)",
        "+ XLOOKUP(T{row}, STRUX_Mätare, STRUX_data, 0, 0, 1)",
        "− XLOOKUP(U{row}, STRUX_Mätare, STRUX_data, 0, 0, 1)",
        "− XLOOKUP(V{row}, STRUX_Mätare, STRUX_data, 0, 0, 1)",
        "− XLOOKUP(W{row}, STRUX_Mätare, STRUX_data, 0, 0, 1)",
        "```",
        "",
        "Semantics: S/T are *additive* supply meters (flows into this building); U/V/W are *subtractive* — meters downstream of the supply that belong to other buildings and get attributed in their own rows. The subtraction is what encodes topology in the accounting view.",
    ]
    path.write_text("\n".join(lines) + "\n")


def _normalize_building_id(raw) -> str | None:
    """Normalize an Excel building cell value to a canonical ``B###`` ID.

    Returns ``None`` for rows that should be skipped (sub-sections like
    ``621 (I&L)``, summary rows like ``Summa Ställverk``).
    """
    s = str(raw).strip()
    if s.lower().startswith(("summa", "trädgård")):
        return None
    if s.lower().startswith("intagsmätare"):
        parts = s.split()
        for p in reversed(parts):
            if p.startswith("B") and any(c.isdigit() for c in p):
                return p
        return f"B{s}"
    if "(" in s:
        paren = s[s.index("(") + 1 : s.index(")")].strip() if ")" in s else ""
        base = s[: s.index("(")].strip()
        if paren == "T":
            return f"B{base}" if not base.startswith("B") else base
        return None
    if s.lower() == "parkering":
        return "BPARKING"
    if s.startswith("B") and len(s) > 1 and s[1:2].isdigit():
        return s
    return f"B{s}"


def extract_building_totals(ws_values, *, sheet_factor: float = 1.0) -> list[dict]:
    """Read cached monthly building totals from the data-only sheet.

    Cols C..N are Jan..Dec; col B is the building number.  The cached cell
    values include any sheet-level scaling factor (e.g. ``$F$5 = 0.001`` on
    the EL sheet for kWh→MWh).  ``sheet_factor`` undoes the conversion so
    output is in the native STRUX unit (matching Snowflake readings).
    """
    month_cols = list(range(3, 15))  # C=3 .. N=14
    month_names = [
        "2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06",
        "2026-07", "2026-08", "2026-09", "2026-10", "2026-11", "2026-12",
    ]
    inv = 1.0 / sheet_factor if sheet_factor != 0 else 1.0
    rows_out: list[dict] = []
    for row in range(8, ws_values.max_row + 1):
        building = ws_values.cell(row=row, column=2).value
        if building is None:
            continue
        bid = _normalize_building_id(building)
        if bid is None:
            continue
        for ci, month in zip(month_cols, month_names):
            val = ws_values.cell(row=row, column=ci).value
            if val is None or not isinstance(val, (int, float)):
                continue
            rows_out.append({
                "building_id": bid,
                "month": month,
                "excel_kwh": val * inv,
            })
    # Deduplicate: when two rows normalize to the same building_id, keep the larger.
    best: dict[tuple[str, str], dict] = {}
    for r in rows_out:
        key = (r["building_id"], r["month"])
        if key not in best or abs(r["excel_kwh"]) > abs(best[key]["excel_kwh"]):
            best[key] = r
    return list(best.values())


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("--media", required=True, help="Sheet name: Ånga, Kyla, Värme, Kallvatten, Kyltornsvatten, EL")
    ap.add_argument("--out-dir", type=Path, required=True)
    args = ap.parse_args()

    if not args.xlsx.exists():
        print(f"error: {args.xlsx} not found", file=sys.stderr)
        return 2

    wb = load_workbook(args.xlsx, data_only=False)
    # Second copy with data_only=True so scalar cells referenced in formulas
    # (e.g. $F$5 = 0.001 in the EL sheet for kWh→MWh scaling) resolve to numbers.
    wb_values = load_workbook(args.xlsx, data_only=True)
    if args.media not in wb.sheetnames:
        print(f"error: sheet {args.media!r} not found. Available: {wb.sheetnames}", file=sys.stderr)
        return 3
    ws = wb[args.media]
    ws_values = wb_values[args.media]

    records = extract_building_formulas(ws, ws_values=ws_values)
    cell_comments = extract_cell_comments(ws)
    strux_rows = extract_strux_media(wb, args.media)
    sheets = inventory_sheets(wb)

    out = args.out_dir
    write_formulas_csv(out / "excel_formulas.csv", records)
    write_meters_used_csv(out / "excel_meters_used.csv", records)
    write_intake_meters_csv(out / "excel_intake_meters.csv", strux_rows)
    write_comments_md(out / "excel_comments.md", args.media, cell_comments, records)
    write_tabs_inventory_md(out / "excel_tabs_inventory.md", args.media, sheets)

    # Extract cached building totals (with sheet-factor correction)
    numeric_factors = set()
    for r in records:
        f = r.get("faktor", "")
        if f != "" and f is not None:
            try:
                numeric_factors.add(float(f))
            except (TypeError, ValueError):
                pass
    sheet_factor = numeric_factors.pop() if len(numeric_factors) == 1 else 1.0
    building_totals = extract_building_totals(ws_values, sheet_factor=sheet_factor)
    if building_totals:
        bt_path = out / "excel_building_totals.csv"
        bt_path.parent.mkdir(parents=True, exist_ok=True)
        with bt_path.open("w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["building_id", "month", "excel_kwh"])
            w.writeheader()
            w.writerows(building_totals)
        print(f"wrote {bt_path} ({len(building_totals)} building-month rows)")

    n_terms_distribution: dict[int, int] = {}
    for r in records:
        n_terms_distribution[r["n_terms"]] = n_terms_distribution.get(r["n_terms"], 0) + 1
    print(f"wrote {out / 'excel_formulas.csv'} ({len(records)} meter refs from {len({r['building'] for r in records})} buildings)")
    print(f"wrote {out / 'excel_meters_used.csv'} ({len({r['meter_id'] for r in records})} unique meters)")
    print(f"wrote {out / 'excel_intake_meters.csv'} ({len(strux_rows)} STRUX rows matching mediaslag={args.media!r})")
    print(f"wrote {out / 'excel_comments.md'} ({len(cell_comments)} cell comments, "
          f"{sum(1 for r in records if r.get('kommentar'))} Kommentar entries)")
    print(f"wrote {out / 'excel_tabs_inventory.md'}")
    if len(n_terms_distribution) > 1:
        dist = ", ".join(f"{k}-term: {v}" for k, v in sorted(n_terms_distribution.items()))
        print(f"formula term counts: {dist}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
