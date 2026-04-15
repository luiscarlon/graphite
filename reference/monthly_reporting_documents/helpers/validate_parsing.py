#!/usr/bin/env python3
"""Validate parsed meter tables against Excel formula calculations.

Usage:
    python validate_parsing.py gtn.xlsx EL
    python validate_parsing.py gtn.xlsx Kyla
    python validate_parsing.py gtn.xlsx  # all tabs
"""

import sys
import re
import openpyxl
from pathlib import Path

SOURCE_DIR = Path(__file__).parent.parent / "source_files"

# Column layout shared across tabs
COL_LETTERS = ['S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA']
COL_NUMS = [19, 20, 21, 22, 23, 24, 25, 26, 27]

# Reference month: February = column D in tabs, column I in STRUX
REF_MONTH_TAB_COL = 4   # D
REF_MONTH_STRUX_COL = 9  # I
REF_MONTH_LABEL = "Feb"


def load_strux(wb_data):
    """Load all STRUX readings for the reference month, keyed by (media, meter)."""
    ws = wb_data['STRUX']
    data = {}
    for row in range(3, ws.max_row + 1):
        media = ws.cell(row=row, column=2).value
        meter = ws.cell(row=row, column=4).value
        val = ws.cell(row=row, column=REF_MONTH_STRUX_COL).value
        if meter and media:
            key = str(meter).strip()
            data[(str(media).strip(), key)] = float(val) if val else 0.0
    return data


def parse_tab_formulas(wb_formulas, tab_name):
    """Parse building formulas from a tab. Returns dict of building → [(meter, sign, coeff)]."""
    ws = wb_formulas[tab_name]

    # Determine column range (EL uses S-AA, Kyla uses S-X)
    max_col_idx = min(len(COL_LETTERS), ws.max_column - COL_NUMS[0] + 1)
    col_letters = COL_LETTERS[:max_col_idx]
    col_nums = COL_NUMS[:max_col_idx]

    formulas = {}
    for row_num in range(8, ws.max_row + 1):
        b_val = ws.cell(row=row_num, column=2).value
        if b_val is None:
            continue
        building = str(b_val).strip()

        r_val = ws.cell(row=row_num, column=18).value  # R = Faktor

        c_cell = ws.cell(row=row_num, column=3)
        formula_text = ""
        if hasattr(c_cell.value, 'text'):
            formula_text = c_cell.value.text
        elif isinstance(c_cell.value, str):
            formula_text = c_cell.value
        if not formula_text or 'XLOOKUP' not in formula_text:
            continue

        # Get meter IDs from columns
        meter_vals = {}
        for col_l, col_n in zip(col_letters, col_nums):
            v = ws.cell(row=row_num, column=col_n).value
            if v:
                meter_vals[col_l] = str(v).strip()

        # Parse each XLOOKUP: extract sign, pre-coefficient, column ref, post-multiplier
        terms = []
        for m in re.finditer(
            r'([=+\-])?\s*(?:\(\$R\d+\*|\(?R\d+\*|\()?(?:([\d.]+)\*)?'
            r'_xlfn\.XLOOKUP\(([A-Z]{1,2})\d+[^)]*\)'
            r'(?:\s*[*/]\s*[\d.]+)*',  # capture post-XLOOKUP multipliers (*/)
            formula_text
        ):
            sign_char = m.group(1) or '+'
            if sign_char == '=':
                sign_char = '+'
            hardcoded = float(m.group(2)) if m.group(2) else None
            col_ref = m.group(3)
            full = m.group(0)
            uses_R = bool(re.search(r'\$?R\d+\*', full))
            is_anchor = '$B$8' in formula_text[max(0, m.start() - 5):m.end() + 50]

            meter = meter_vals.get(col_ref, '')
            if not meter:
                continue

            if uses_R and r_val:
                coeff = float(r_val)
            elif hardcoded:
                coeff = hardcoded
            else:
                coeff = 1.0

            # Post-XLOOKUP multipliers (e.g. *24*31/1000)
            post = full[full.rfind(')') + 1:]
            if post:
                post_mult = 1.0
                for op in re.finditer(r'([*/])\s*([\d.]+)', post):
                    if op.group(1) == '*':
                        post_mult *= float(op.group(2))
                    else:
                        post_mult /= float(op.group(2))
                coeff *= post_mult

            terms.append((meter, sign_char, coeff, is_anchor))

        if terms:
            formulas[building] = terms

    return formulas


def get_excel_values(wb_data, tab_name):
    """Get cached building values from the Excel for the reference month."""
    ws = wb_data[tab_name]
    values = {}
    for row_num in range(8, ws.max_row + 1):
        b_val = ws.cell(row=row_num, column=2).value
        if b_val is None:
            continue
        building = str(b_val).strip()
        val = ws.cell(row=row_num, column=REF_MONTH_TAB_COL).value
        if val is not None:
            values[building] = float(val)
    return values


def get_unit_factor(wb_data, tab_name):
    """Get unit factor from F5."""
    ws = wb_data[tab_name]
    val = ws.cell(row=5, column=6).value
    return float(val) if val else 1.0


def media_for_tab(tab_name):
    """Map tab name to STRUX media filter."""
    mapping = {
        'EL': 'El',
        'Kyla': 'Kyla',
        'Värme': 'Värme',
        'Ånga': 'Ånga',
        'Kallvatten': 'Kallvatten',
        'Kyltornsvatten': 'Kyltornsvatten',
    }
    return mapping.get(tab_name)


def validate_tab(source_file, tab_name):
    """Validate a single tab. Returns (matches, mismatches, details)."""
    path = SOURCE_DIR / source_file
    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    wb_data = openpyxl.load_workbook(path, data_only=True)

    strux = load_strux(wb_data)
    formulas = parse_tab_formulas(wb_formulas, tab_name)
    excel_vals = get_excel_values(wb_data, tab_name)
    uf = get_unit_factor(wb_data, tab_name)
    media = media_for_tab(tab_name)

    def strux_val(meter):
        """Look up a meter in STRUX. Try tab-specific media first, then any."""
        if media:
            val = strux.get((media, meter))
            if val is not None:
                return val
        # Fallback: search all media
        for (m, mid), v in strux.items():
            if mid == meter:
                return v
        return 0.0

    # For Kyla: compute B600-KB2 first if needed
    computed_meters = {}
    if 'B600-KB2' in formulas:
        b600_terms = formulas['B600-KB2']
        b600_val = 0.0
        for meter, sign, coeff, _ in b600_terms:
            v = strux_val(meter)
            b600_val += (1 if sign == '+' else -1) * coeff * v
        computed_meters['B600-KB2'] = b600_val

    matches = 0
    mismatches = 0
    details = []

    for building, terms in formulas.items():
        excel_val = excel_vals.get(building)
        if excel_val is None or excel_val == 0:
            continue

        calc = 0.0
        for meter, sign, coeff, is_anchor in terms:
            if is_anchor and meter in computed_meters:
                v = computed_meters[meter]
            else:
                v = strux_val(meter)
            calc += (1 if sign == '+' else -1) * coeff * v
        calc *= uf

        diff = abs(excel_val - calc)
        ok = diff < 0.001
        if ok:
            matches += 1
        else:
            mismatches += 1

        details.append({
            'building': building,
            'excel': excel_val,
            'calc': calc,
            'diff': diff,
            'ok': ok,
        })

    return matches, mismatches, details


def main():
    if len(sys.argv) < 2:
        print("Usage: python validate_parsing.py <source_file> [tab_name]")
        sys.exit(1)

    source_file = sys.argv[1]
    tab_names = [sys.argv[2]] if len(sys.argv) > 2 else None

    if tab_names is None:
        # Discover tabs
        path = SOURCE_DIR / source_file
        wb = openpyxl.load_workbook(path, read_only=True)
        known_tabs = {'EL', 'Kyla', 'Värme', 'Ånga', 'Kallvatten', 'Kyltornsvatten'}
        tab_names = [s for s in wb.sheetnames if s in known_tabs]
        wb.close()

    for tab_name in tab_names:
        print(f"\n{'=' * 60}")
        print(f"  {source_file} — {tab_name} ({REF_MONTH_LABEL})")
        print(f"{'=' * 60}")

        matches, mismatches, details = validate_tab(source_file, tab_name)

        for d in details:
            status = '✓' if d['ok'] else f"✗ {d['diff']:.4f}"
            print(f"  {d['building']:<20} {d['excel']:<12.4f} {d['calc']:<12.4f} {status}")

        print(f"\n  Result: {matches} match, {mismatches} mismatch")


if __name__ == '__main__':
    main()
